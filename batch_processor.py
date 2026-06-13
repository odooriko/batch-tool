import os
import re
import openpyxl
from openpyxl import Workbook
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

# ==================== 原有功能函数（保持不变） ====================
res = ['临时供配电工程','水电一标段','水电一标','水电二标段','水电二标','水电三标','水电四标','消防一标段','消防二标','通风安装工程','弱电一标','弱电二标']

def get_adjusted_values_from_folder(folder_path):
    """遍历文件夹，提取每个文件中'合计'和'应扣除临设分摊费'所在行的Q列值"""
    result = {}
    for filename in os.listdir(folder_path):
        if not filename.endswith('.xlsx'):
            continue
        pattern = r'20\d{2}年\d{1,2}月'
        match = re.search(pattern, filename)
        if match:
            month = match.group(0)
        else:
            raise ValueError(f"文件名 '{filename}' 中未找到规范的月份格式（应为 20XX年X月 或 20XX年XX月）")

        key = None
        for item in res:
            if item in filename:
                key = item
                break
        if key is None:
            raise ValueError(f"错误：文件 {filename} 中未匹配到 res 列表中的任何项目，请检查文件名。")

        file_path = os.path.join(folder_path, filename)
        wb = openpyxl.load_workbook(file_path, data_only=True)
        values = {"合计": None, "应扣除临设分摊费": None}

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            found_合计 = False
            found_分摊 = False
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value == "合计" and not found_合计:
                        values["合计"] = ws[f"Q{cell.row}"].value
                        found_合计 = True
                    elif cell.value == "应扣除临设分摊费" and not found_分摊:
                        values["应扣除临设分摊费"] = -ws[f"Q{cell.row}"].value
                        found_分摊 = True
                    if found_合计 and found_分摊:
                        break
                if found_合计 and found_分摊:
                    break
        result[key + "-" + month] = values
    return result

def replace_month_in_filename(filename, new_month_str):
    """将文件名中的“2026年X月”替换为 new_month_str。如果没有匹配，则在文件名开头加上 new_month_str。"""
    pattern = r'2026年\d+月'
    if re.search(pattern, filename):
        return re.sub(pattern, new_month_str, filename)
    else:
        return new_month_str + filename

def modify_new_file_with_values(original_file_path, output_dict, folder_name, save_path=None):
    """
    修改目标文件（仅处理工作表名为“进度批量”）：
       - 找到单元格值包含 output_dict 键的项目名称（即 key.split('-')[0]）的行作为合计行，下一行为应扣除行。
       - 合计行：J列替换为合计值，M列和P列增加合计值。
       - 应扣除行：J列替换为应扣除值，M列和P列增加应扣除值。
       - 将第一个工作表的A1单元格的前7个字符替换为 folder_name。
       - 输出文件名：将原文件名中的月份部分替换为 folder_name。
    """
    if save_path is None:
        dir_name = os.path.dirname(original_file_path)
        base_name = os.path.basename(original_file_path)
        new_base_name = replace_month_in_filename(base_name, folder_name)
        save_path = os.path.join(dir_name, new_base_name)

    wb = openpyxl.load_workbook(original_file_path, data_only=False)

    # 构建项目名称（不含月份）到值的映射
    project_map = {}
    for key, values in output_dict.items():
        project_name = key.split('-')[0]
        project_map[project_name] = values

    # 修改第一个工作表的 A1 单元格
    first_sheet = wb[wb.sheetnames[0]]
    a1_cell = first_sheet["A1"]
    original_value = a1_cell.value
    if original_value is not None:
        original_str = str(original_value)
        if len(original_str) >= 7:
            new_value = folder_name + original_str[7:]
        else:
            new_value = folder_name
        a1_cell.value = new_value
    else:
        a1_cell.value = folder_name

    # 只处理名为“进度批量”的工作表
    target_sheet_name = "进度批量"
    if target_sheet_name not in wb.sheetnames:
        raise ValueError(f"工作簿中不存在名为 '{target_sheet_name}' 的工作表，请检查。")

    ws = wb[target_sheet_name]

    processed = set()
    total_projects = len(project_map)

    for row in ws.iter_rows():
        for cell in row:
            if len(processed) == total_projects:
                break
            cell_value = cell.value
            if not isinstance(cell_value, str):
                continue
            matched_project = None
            for proj_name in project_map.keys():
                if proj_name in cell_value and proj_name not in processed:
                    matched_project = proj_name
                    break
            if matched_project is None:
                continue

            processed.add(matched_project)
            values = project_map[matched_project]
            row_num = cell.row

            # 合计行操作
            sum_val = values["合计"]
            if sum_val is not None:
                ws[f"J{row_num}"].value = sum_val
                m_cell = ws[f"M{row_num}"]
                orig_m = m_cell.value if m_cell.value is not None else 0
                try:
                    orig_m = float(orig_m)
                except (ValueError, TypeError):
                    orig_m = 0.0
                m_cell.value = orig_m + sum_val
                p_cell = ws[f"P{row_num}"]
                orig_p = p_cell.value if p_cell.value is not None else 0
                try:
                    orig_p = float(orig_p)
                except (ValueError, TypeError):
                    orig_p = 0.0
                p_cell.value = orig_p + sum_val

            # 应扣除行操作（下一行）
            deduct_val = values["应扣除临设分摊费"]
            if deduct_val is not None:
                row_deduct = row_num + 1
                ws[f"J{row_deduct}"].value = deduct_val
                m_cell2 = ws[f"M{row_deduct}"]
                orig_m2 = m_cell2.value if m_cell2.value is not None else 0
                try:
                    orig_m2 = float(orig_m2)
                except (ValueError, TypeError):
                    orig_m2 = 0.0
                m_cell2.value = orig_m2 + deduct_val
                p_cell2 = ws[f"P{row_deduct}"]
                orig_p2 = p_cell2.value if p_cell2.value is not None else 0
                try:
                    orig_p2 = float(orig_p2)
                except (ValueError, TypeError):
                    orig_p2 = 0.0
                p_cell2.value = orig_p2 + deduct_val

        if len(processed) == total_projects:
            break

    wb.save(save_path)
    wb.close()
    return save_path

def save_result_to_excel(result_dict, folder_path, filename="分包表统计结果.xlsx"):
    """将结果字典保存为 Excel 文件，保存到 folder_path 的父目录下"""
    parent_dir = os.path.dirname(folder_path)
    output_path = os.path.join(parent_dir, filename)
    wb = Workbook()
    ws = wb.active
    ws.title = "汇总数据"
    ws.append(["项目名称", "合计", "应扣除临设分摊费"])
    for key, values in result_dict.items():
        ws.append([key, values["合计"], values["应扣除临设分摊费"]])
    wb.save(output_path)
    return output_path


# ==================== GUI 界面 ====================
class Application:
    def __init__(self, master):
        self.master = master
        master.title("小洲作弊器")
        master.geometry("700x400")
        master.resizable(False, False)

        # 全局缓存统计结果
        self.cached_output = None
        self.last_folder = None

        # 样式
        style = ttk.Style()
        style.configure("TButton", font=("微软雅黑", 10), padding=6)
        style.configure("TLabel", font=("微软雅黑", 10))
        style.configure("TEntry", font=("微软雅黑", 10))

        # 标题
        title_label = ttk.Label(master, text="进度批量表生成工具", font=("微软雅黑", 16, "bold"))
        title_label.grid(row=0, column=0, columnspan=3, pady=15)

        # 说明文字
        desc_label = ttk.Label(master, text="请选择分包表文件夹路径（命名规范如 2026年6月，其他月同理）", foreground="gray")
        desc_label.grid(row=1, column=0, columnspan=3, pady=(0, 10), sticky="w", padx=20)

        # 分包表文件夹选择
        ttk.Label(master, text="分包表路径：").grid(row=2, column=0, padx=20, pady=5, sticky="e")
        self.folder_var = tk.StringVar()
        self.folder_entry = ttk.Entry(master, textvariable=self.folder_var, width=60, font=("微软雅黑", 9))
        self.folder_entry.grid(row=2, column=1, padx=5, pady=5)
        self.folder_btn = ttk.Button(master, text="浏览", command=self.select_folder)
        self.folder_btn.grid(row=2, column=2, padx=10, pady=5)

        # 上月批量表文件选择
        ttk.Label(master, text="上月批量进度表路径：").grid(row=3, column=0, padx=20, pady=5, sticky="e")
        self.file_var = tk.StringVar()
        self.file_entry = ttk.Entry(master, textvariable=self.file_var, width=60, font=("微软雅黑", 9))
        self.file_entry.grid(row=3, column=1, padx=5, pady=5)
        self.file_btn = ttk.Button(master, text="浏览", command=self.select_file)
        self.file_btn.grid(row=3, column=2, padx=10, pady=5)

        # 按钮框架
        btn_frame = ttk.Frame(master)
        btn_frame.grid(row=4, column=0, columnspan=3, pady=30)

        self.stat_btn = ttk.Button(btn_frame, text="分包表结果统计", command=self.do_statistics, width=20)
        self.stat_btn.pack(side="left", padx=20)

        self.gen_btn = ttk.Button(btn_frame, text="生成本月进度批量表", command=self.do_generate, width=20)
        self.gen_btn.pack(side="right", padx=20)

        # 状态栏
        self.status_var = tk.StringVar()
        self.status_var.set("就绪")
        self.status_bar = ttk.Label(master, textvariable=self.status_var, relief="sunken", anchor="w", font=("微软雅黑", 9))
        self.status_bar.grid(row=5, column=0, columnspan=3, sticky="ew", padx=10, pady=(10, 5))

    def select_folder(self):
        folder_selected = filedialog.askdirectory(title="选择分包表文件夹")
        if folder_selected:
            self.folder_var.set(folder_selected)
            self.status_var.set("已选择分包表文件夹")

    def select_file(self):
        file_selected = filedialog.askopenfilename(title="选择上月进度批量表", filetypes=[("Excel文件", "*.xlsx")])
        if file_selected:
            self.file_var.set(file_selected)
            self.status_var.set("已选择上月批量表文件")

    def do_statistics(self):
        folder_path = self.folder_var.get().strip()
        if not folder_path:
            messagebox.showerror("错误", "请先选择分包表文件夹路径")
            return
        if not os.path.isdir(folder_path):
            messagebox.showerror("错误", "分包表文件夹路径无效")
            return

        try:
            self.status_var.set("正在统计分包表，请稍候...")
            self.master.update()

            # 执行统计
            output = get_adjusted_values_from_folder(folder_path)

            # 保存结果
            out_file = save_result_to_excel(output, folder_path)

            # 缓存结果
            self.cached_output = output
            self.last_folder = folder_path

            messagebox.showinfo("完成", f"统计完成！\n结果已保存至：{out_file}")
            self.status_var.set("统计完成")
        except Exception as e:
            messagebox.showerror("统计失败", str(e))
            self.status_var.set("统计失败，请检查错误信息")

    def do_generate(self):
        # 检查是否已有统计缓存
        if self.cached_output is None:
            # 尝试直接统计（如果文件夹已选且未统计过）
            folder_path = self.folder_var.get().strip()
            if not folder_path or not os.path.isdir(folder_path):
                messagebox.showerror("错误", "请先选择有效的分包表文件夹，并点击“分包表结果统计”")
                return
            # 自动执行统计
            try:
                self.status_var.set("未找到统计缓存，正在自动统计...")
                self.master.update()
                output = get_adjusted_values_from_folder(folder_path)
                self.cached_output = output
                self.last_folder = folder_path
            except Exception as e:
                messagebox.showerror("统计失败", f"自动统计失败：{str(e)}\n请手动点击“分包表结果统计”按钮")
                self.status_var.set("统计失败")
                return

        # 检查上月文件路径
        src_file = self.file_var.get().strip()
        if not src_file:
            messagebox.showerror("错误", "请先选择上月进度批量表文件")
            return
        if not os.path.isfile(src_file):
            messagebox.showerror("错误", "上月进度批量表文件不存在")
            return

        try:
            folder_name = os.path.basename(self.last_folder)  # 如 "2026年6月"
            self.status_var.set("正在生成当月进度批量表，请稍候...")
            self.master.update()

            new_file = modify_new_file_with_values(src_file, self.cached_output, folder_name)

            messagebox.showinfo("完成", f"生成成功！\n新文件已保存至：{new_file}")
            self.status_var.set("生成成功")
        except Exception as e:
            messagebox.showerror("生成失败", str(e))
            self.status_var.set("生成失败，请检查错误信息")


if __name__ == "__main__":
    root = tk.Tk()
    app = Application(root)
    root.mainloop()